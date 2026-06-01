"""从 xlsx/csv 文件加载 IMU 加速度数据"""

import csv
import numpy as np
import openpyxl
from pathlib import Path


def load_xlsx(path: str | Path) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    读取 xlsx 中的 3 轴加速度数据。
    假设列布局: PC Date, PC Time, Line 1(X), Line 2(Y), Line 3(Z)
    返回 (ax, ay, az), dtype=float32
    """
    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    data = rows[1:]  # 跳过表头
    n = len(data)
    ax = np.empty(n, dtype=np.float32)
    ay = np.empty(n, dtype=np.float32)
    az = np.empty(n, dtype=np.float32)
    for i, row in enumerate(data):
        ax[i] = row[2]
        ay[i] = row[3]
        az[i] = row[4]
    wb.close()
    return ax, ay, az


def load_csv(path: str | Path) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    读取 tcp_receiver.py 生成的 CSV 文件。
    列布局: timestamp_us, datetime, acc_x, acc_y, acc_z
    返回 (ax, ay, az), dtype=float32
    """
    ax_list, ay_list, az_list = [], [], []
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            ax_list.append(float(row['acc_x']))
            ay_list.append(float(row['acc_y']))
            az_list.append(float(row['acc_z']))
    return (np.array(ax_list, dtype=np.float32),
            np.array(ay_list, dtype=np.float32),
            np.array(az_list, dtype=np.float32))


def remove_dc_offset(
    ax: np.ndarray, ay: np.ndarray, az: np.ndarray, static_n: int = 10000
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    用前 static_n 行静态数据计算直流偏移, 减去后返回交流分量。
    同时打印偏移值供参考。
    """
    n = min(static_n, len(ax))
    dc_x = ax[:n].mean()
    dc_y = ay[:n].mean()
    dc_z = az[:n].mean()
    print(f"直流偏移 (前{n}行): X={dc_x:.6f}, Y={dc_y:.6f}, Z={dc_z:.6f}")
    return ax - dc_x, ay - dc_y, az - dc_z


def find_data_file(data_dir: str | Path) -> Path | None:
    """在目录中找到第一个 csv 或 xlsx 文件 (优先 csv)"""
    p = Path(data_dir)
    csv_files = list(p.glob("*.csv"))
    if csv_files:
        return csv_files[0]
    xlsx_files = list(p.glob("*.xlsx"))
    if xlsx_files:
        return xlsx_files[0]
    return None


# 保留旧函数名作为别名
def find_xlsx(data_dir: str | Path) -> Path | None:
    """兼容旧接口"""
    return find_data_file(data_dir)
